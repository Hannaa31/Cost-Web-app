import sys, os, io
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from typing import List, Optional
from datetime import datetime
from fastapi import APIRouter, Depends, HTTPException, status
from fastapi.responses import StreamingResponse
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session
from database import get_db
import models, schemas, auth

router = APIRouter(prefix="/api/projects", tags=["Projects & Workspace"])

@router.get("", response_model=List[schemas.ProjectResponse])
def list_user_projects(
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    # Admins see all projects or user sees their own projects
    if current_user.role == models.UserRole.admin:
        projects = db.query(models.Project).order_by(models.Project.created_at.desc()).all()
    else:
        projects = db.query(models.Project).filter(models.Project.created_by == current_user.id).order_by(models.Project.created_at.desc()).all()
    
    dirty = False
    for p in projects:
        for item in p.line_items:
            margin_pct = item.selected_rate.margin_pct if (item.selected_rate and hasattr(item.selected_rate, 'margin_pct') and item.selected_rate.margin_pct is not None) else (p.global_margin_pct or 0.0)
            expected_cost = round((item.calculated_escalated_rate * item.quantity) * (1.0 + margin_pct), 2)
            if item.total_item_cost != expected_cost:
                item.total_item_cost = expected_cost
                dirty = True
    if dirty:
        db.commit()
    return projects

@router.post("", response_model=schemas.ProjectResponse, status_code=status.HTTP_201_CREATED)
def create_project(
    project_in: schemas.ProjectCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    new_project = models.Project(
        created_by=current_user.id,
        name=project_in.name,
        client=project_in.client,
        global_margin_pct=project_in.global_margin_pct,
        global_erection_pct=project_in.global_erection_pct,
        default_annual_escalation_pct=project_in.default_annual_escalation_pct,
        conveyor_length_mtr=project_in.conveyor_length_mtr,
        total_mine_life_years=project_in.total_mine_life_years,
        phases=project_in.phases
    )
    db.add(new_project)
    db.commit()
    db.refresh(new_project)
    return new_project

@router.get("/{project_id}", response_model=schemas.ProjectResponse)
def get_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != models.UserRole.admin and project.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to access this workspace")
    for item in project.line_items:
        margin_pct = item.selected_rate.margin_pct if (item.selected_rate and hasattr(item.selected_rate, 'margin_pct') and item.selected_rate.margin_pct is not None) else (project.global_margin_pct or 0.0)
        expected_cost = round((item.calculated_escalated_rate * item.quantity) * (1.0 + margin_pct), 2)
        if item.total_item_cost != expected_cost:
            item.total_item_cost = expected_cost
    db.commit()
    return project

@router.put("/{project_id}", response_model=schemas.ProjectResponse)
def update_project(
    project_id: int,
    project_update: schemas.ProjectUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != models.UserRole.admin and project.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to modify this workspace")
    
    if project_update.name is not None:
        project.name = project_update.name
    if project_update.client is not None:
        project.client = project_update.client
    if project_update.global_margin_pct is not None:
        project.global_margin_pct = project_update.global_margin_pct
    if project_update.global_erection_pct is not None:
        project.global_erection_pct = project_update.global_erection_pct
    if project_update.conveyor_length_mtr is not None:
        project.conveyor_length_mtr = project_update.conveyor_length_mtr
    if project_update.total_mine_life_years is not None:
        project.total_mine_life_years = project_update.total_mine_life_years
    if project_update.phases is not None:
        project.phases = project_update.phases
    
    if project_update.default_annual_escalation_pct is not None and project_update.default_annual_escalation_pct != project.default_annual_escalation_pct:
        project.default_annual_escalation_pct = project_update.default_annual_escalation_pct
        line_items = db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id).all()
        now = datetime.utcnow()
        for item in line_items:
            rate = db.query(models.MasterRate).filter(models.MasterRate.id == item.selected_rate_id).first()
            if rate:
                years_elapsed = max(0.0, (now - rate.quotation_date).days / 365.25)
                multiplier = (1 + project.default_annual_escalation_pct) ** years_elapsed
                item.calculated_escalated_rate = round(rate.base_rate * multiplier, 2)
                item_margin_pct = rate.margin_pct if (rate and hasattr(rate, 'margin_pct') and rate.margin_pct is not None) else (project.global_margin_pct or 0.0)
                item.total_item_cost = round((item.calculated_escalated_rate * item.quantity) * (1.0 + item_margin_pct), 2)
    elif project_update.default_annual_escalation_pct is not None:
        project.default_annual_escalation_pct = project_update.default_annual_escalation_pct

    db.commit()
    db.refresh(project)
    return project

@router.delete("/{project_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_project(
    project_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    if current_user.role != models.UserRole.admin and project.created_by != current_user.id:
        raise HTTPException(status_code=403, detail="Not authorized to delete this workspace")
    
    # Delete associated line items first
    db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id).delete()
    db.delete(project)
    db.commit()
    return None

@router.get("/{project_id}/line-items", response_model=List[schemas.EstimateLineItemResponse])
def get_project_line_items(
    project_id: int,
    domain: Optional[models.DomainType] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    query = db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id)
    if domain:
        query = query.filter(models.EstimateLineItem.domain == domain)
    return query.all()

@router.post("/{project_id}/line-items", response_model=schemas.EstimateLineItemResponse, status_code=status.HTTP_201_CREATED)
def add_line_item(
    project_id: int,
    item_in: schemas.EstimateLineItemCreate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    rate = db.query(models.MasterRate).filter(models.MasterRate.id == item_in.selected_rate_id).first()
    if not rate:
        raise HTTPException(status_code=404, detail="Selected Master Rate not found")
    
    cat = db.query(models.EquipmentCategory).filter(models.EquipmentCategory.id == item_in.category_id).first()
    item_domain = item_in.domain or (cat.domain if cat else models.DomainType.Mechanical)
    parent_val = item_in.parent_category or (cat.parent_category if cat else None)

    # Calculate escalated rate using project default escalation
    now = datetime.utcnow()
    years_elapsed = max(0.0, (now - rate.quotation_date).days / 365.25)
    multiplier = (1 + project.default_annual_escalation_pct) ** years_elapsed
    escalated_rate = round(rate.base_rate * multiplier, 2)
    base_cost = round(escalated_rate * item_in.quantity, 2)
    item_margin_pct = rate.margin_pct if (rate and hasattr(rate, 'margin_pct') and rate.margin_pct is not None) else (project.global_margin_pct or 0.0)
    total_cost = round(base_cost * (1.0 + item_margin_pct), 2)

    new_item = models.EstimateLineItem(
        project_id=project_id,
        category_id=item_in.category_id,
        selected_rate_id=item_in.selected_rate_id,
        domain=item_domain,
        parent_category=parent_val,
        phase_name=item_in.phase_name or "Phase 1",
        quantity=item_in.quantity,
        calculated_escalated_rate=escalated_rate,
        total_item_cost=total_cost
    )
    db.add(new_item)
    db.commit()
    db.refresh(new_item)
    return new_item

@router.put("/{project_id}/line-items/{item_id}", response_model=schemas.EstimateLineItemResponse)
def update_line_item(
    project_id: int,
    item_id: int,
    item_update: schemas.EstimateLineItemUpdate,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")

    item = db.query(models.EstimateLineItem).filter(
        models.EstimateLineItem.id == item_id,
        models.EstimateLineItem.project_id == project_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Line item not found")
    
    if item_update.selected_rate_id is not None and item_update.selected_rate_id != item.selected_rate_id:
        new_rate = db.query(models.MasterRate).filter(models.MasterRate.id == item_update.selected_rate_id).first()
        if not new_rate:
            raise HTTPException(status_code=404, detail="New Master Rate not found")
        item.selected_rate_id = new_rate.id
        now = datetime.utcnow()
        years_elapsed = max(0.0, (now - new_rate.quotation_date).days / 365.25)
        multiplier = (1 + project.default_annual_escalation_pct) ** years_elapsed
        item.calculated_escalated_rate = round(new_rate.base_rate * multiplier, 2)
    
    if item_update.quantity is not None:
        item.quantity = item_update.quantity
    if item_update.phase_name is not None:
        item.phase_name = item_update.phase_name

    item_margin_pct = item.selected_rate.margin_pct if (item.selected_rate and hasattr(item.selected_rate, 'margin_pct') and item.selected_rate.margin_pct is not None) else (project.global_margin_pct or 0.0)
    item.total_item_cost = round((item.calculated_escalated_rate * item.quantity) * (1.0 + item_margin_pct), 2)
    db.commit()
    db.refresh(item)
    return item

@router.delete("/{project_id}/line-items/{item_id}", status_code=status.HTTP_204_NO_CONTENT)
def delete_line_item(
    project_id: int,
    item_id: int,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    item = db.query(models.EstimateLineItem).filter(
        models.EstimateLineItem.id == item_id,
        models.EstimateLineItem.project_id == project_id
    ).first()
    if not item:
        raise HTTPException(status_code=404, detail="Line item not found")
    db.delete(item)
    db.commit()
    return None

def get_item_uom(category_name: str) -> str:
    n = category_name.lower()
    if any(k in n for k in ['belt', 'structure', 'cable', 'wire mesh', 'hood', 'piping', 'railing']):
        return "Mtr"
    elif any(k in n for k in ['chute', 'steelwork', 'plate', 'workshop', 'shifting', 'contingency', 'package']):
        return "Lot"
    elif any(k in n for k in ['foundation', 'concrete', 'excavation']):
        return "Cu.m"
    elif any(k in n for k in ['shed', 'road', 'sheeting']):
        return "Sq.m"
    else:
        return "Set"

def populate_sheet(ws, project, line_items, domain_name=None):
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True)
    header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
    accent_fill = PatternFill(start_color="0284c7" if not domain_name else ("059669" if domain_name == "Mechanical" else ("d97706" if domain_name == "Electrical" else "4f46e5")), end_color="0284c7", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    if domain_name:
        ws.title = domain_name
        items = [i for i in line_items if (i.domain == domain_name or (i.category and i.category.domain == domain_name))]
    else:
        ws.title = "Cost Report"
        items = line_items

    ws.merge_cells("A1:I1")
    ws["A1"] = f"PROJECT ESTIMATION COST REPORT: {project.name.upper()} ({domain_name.upper() if domain_name else 'ALL DOMAINS'})"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws["A3"] = "Client Entity:"
    ws["B3"] = project.client
    ws["A4"] = "Generated Date:"
    ws["B4"] = datetime.now().strftime("%d/%m/%Y")
    ws["D3"] = "EPC Markup:"
    ws["E3"] = f"{project.global_margin_pct * 100:.1f}%"
    ws["D4"] = "Erection:"
    ws["E4"] = f"{project.global_erection_pct * 100:.1f}%"
    for cell in ["A3", "A4", "D3", "D4"]:
        ws[cell].font = bold_font

    headers = ["Item #", "Equipment Category", "Specifications", "Vendor Name", "Base Rate (INR)", "Escalated Rate (INR)", "Remarks", "Quantity", "Total Cost (INR)"]
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col_num, value=h_text)
        cell.font = header_font
        cell.fill = accent_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[6].height = 22

    subtotal = 0.0
    row_idx = 7
    for i, item in enumerate(items, 1):
        cat_name = item.category.name if item.category else f"Category #{item.category_id}"
        vendor_name = item.selected_rate.vendor_name if item.selected_rate else f"Rate #{item.selected_rate_id}"
        base_rate = item.selected_rate.base_rate if item.selected_rate else 0.0
        
        spec_dict = item.selected_rate.specifications if item.selected_rate and item.selected_rate.specifications else {}
        if isinstance(spec_dict, dict) and spec_dict:
            spec_str = ", ".join(f"{k}: {v}" for k, v in spec_dict.items() if k != "Remarks")
        else:
            spec_str = str(spec_dict) if spec_dict else "Standard"
        
        remarks_str = ""
        if item.selected_rate:
            remarks_str = item.selected_rate.remarks or (spec_dict.get("Remarks") if isinstance(spec_dict, dict) else "") or "-"

        ws.cell(row=row_idx, column=1, value=i).alignment = Alignment(horizontal="center")
        ws.cell(row=row_idx, column=2, value=cat_name)
        ws.cell(row=row_idx, column=3, value=spec_str)
        ws.cell(row=row_idx, column=4, value=vendor_name)
        
        base_cell = ws.cell(row=row_idx, column=5, value=base_rate)
        base_cell.number_format = '₹#,##0.00'
        base_cell.alignment = Alignment(horizontal="right")

        rate_cell = ws.cell(row=row_idx, column=6, value=item.calculated_escalated_rate)
        rate_cell.number_format = '₹#,##0.00'
        rate_cell.alignment = Alignment(horizontal="right")
        
        ws.cell(row=row_idx, column=7, value=remarks_str)

        qty_cell = ws.cell(row=row_idx, column=8, value=item.quantity)
        qty_cell.alignment = Alignment(horizontal="right")
        
        if item.quantity > 0 and item.calculated_escalated_rate > 0:
            cost_cell = ws.cell(row=row_idx, column=9, value=f"=F{row_idx}*G{row_idx}")
        else:
            cost_cell = ws.cell(row=row_idx, column=9, value=item.total_item_cost)
        cost_cell.number_format = '₹#,##0.00'
        cost_cell.alignment = Alignment(horizontal="right")
        
        for c in range(1, 10):
            ws.cell(row=row_idx, column=c).border = thin_border
            
        subtotal += item.total_item_cost
        ws.row_dimensions[row_idx].height = 18
        row_idx += 1

    data_end_row = row_idx - 1
    if len(items) == 0:
        ws.merge_cells(f"A{row_idx}:I{row_idx}")
        ws.cell(row=row_idx, column=1, value=f"No {domain_name or ''} equipment added yet.").alignment = Alignment(horizontal="center")
        row_idx += 1

    margin_cost = round(subtotal * project.global_margin_pct, 2)
    erection_cost = round((subtotal + margin_cost) * project.global_erection_pct, 2)
    grand_total = round(subtotal + margin_cost + erection_cost, 2)

    row_idx += 1
    s_sub_row = row_idx
    s_ere_row = row_idx + 1
    s_grd_row = row_idx + 2

    summary_rows = [
        ("Equipment Subtotal:", f"=SUM(I4:I{data_end_row})" if len(items) > 0 else subtotal),
        (f"Erection ({project.global_erection_pct * 100:.1f}%):", f"=I{s_sub_row}*{project.global_erection_pct}"),
        ("GRAND ESTIMATE TOTAL:", f"=SUM(I{s_sub_row}:I{s_ere_row})")
    ]

    for label, val in summary_rows:
        ws.merge_cells(f"G{row_idx}:H{row_idx}")
        label_cell = ws.cell(row=row_idx, column=7, value=label)
        label_cell.font = bold_font if label != "GRAND ESTIMATE TOTAL:" else Font(name="Calibri", size=11, bold=True, color="004B23")
        label_cell.alignment = Alignment(horizontal="right")
        
        val_cell = ws.cell(row=row_idx, column=9, value=val)
        val_cell.font = bold_font if label != "GRAND ESTIMATE TOTAL:" else Font(name="Calibri", size=11, bold=True, color="004B23")
        val_cell.number_format = '₹#,##0.00'
        val_cell.alignment = Alignment(horizontal="right")
        val_cell.border = thin_border
        ws.row_dimensions[row_idx].height = 20 if label == "GRAND ESTIMATE TOTAL:" else 18
        row_idx += 1

    adjust_column_widths(ws)
    return subtotal

def adjust_column_widths(ws, min_width=12, max_width=45):
    # Find all merged ranges that span across more than 1 column
    multi_col_merges = set()
    for mrange in ws.merged_cells.ranges:
        if mrange.max_col > mrange.min_col:
            for r in range(mrange.min_row, mrange.max_row + 1):
                for c in range(mrange.min_col, mrange.max_col + 1):
                    multi_col_merges.add((r, c))

    for col in ws.columns:
        col_idx = col[0].column
        col_letter = get_column_letter(col_idx)
        
        is_blank_col = True
        max_len = 0
        for cell in col:
            val = cell.value
            if val is not None and str(val).strip() != "":
                is_blank_col = False
                if (cell.row, col_idx) in multi_col_merges:
                    continue
                val_str = str(val)
                if isinstance(val, (int, float)) and cell.number_format and '₹' in str(cell.number_format):
                    val_str = f"₹{val:,.2f}"
                max_len = max(max_len, len(val_str))
        
        if is_blank_col:
            ws.column_dimensions[col_letter].width = 4
        elif col_idx == 1:
            ws.column_dimensions[col_letter].width = max(max_len + 3, 8)
        else:
            ws.column_dimensions[col_letter].width = min(max(max_len + 3, min_width), max_width)

def populate_mech_summary_sheet(ws, project, line_items):
    ws.title = "Mech Summary"
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    sub_font = Font(name="Calibri", size=10, bold=True, color="1E293B")
    bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    
    # Lighter and softer color palette
    header_fill = PatternFill(start_color="4F6D7A", end_color="4F6D7A", fill_type="solid") # Soft steel blue
    sub_fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid") # Soft light slate grey
    rust_fill = PatternFill(start_color="D97706", end_color="D97706", fill_type="solid") # Soft warm golden amber
    group_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Soft cool grey banner
    subtotal_fill = PatternFill(start_color="E0F2FE", end_color="E0F2FE", fill_type="solid") # Soft ice blue total banner
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    items = [i for i in line_items if str(i.domain.value if hasattr(i.domain, 'value') else i.domain).lower() == "mechanical" or (i.category and str(i.category.domain.value if hasattr(i.category.domain, 'value') else i.category.domain).lower() == "mechanical")]
    phases = project.phases or []
    max_cols = 9 + len(phases) * 7 if len(phases) > 0 else 9

    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    ws["A1"] = f"PROJECT ESTIMATION COST REPORT: {project.name.upper()} (MECHANICAL SUMMARY)"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    # Row 3 Date cell exactly like screenshot (B3 brown/amber cell)
    ws["B3"] = datetime.now().strftime("%d.%m.%Y")
    ws["B3"].font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    ws["B3"].fill = rust_fill
    ws["B3"].alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=3, column=2).border = thin_border

    # Top Block Header in Row 4 for All Phases
    ws.merge_cells("D4:I4")
    mine_life = project.total_mine_life_years or 26
    ws["D4"] = f"All Phases - 0th year to {mine_life}th year"
    ws["D4"].font = header_font
    ws["D4"].fill = header_fill
    ws["D4"].alignment = Alignment(horizontal="center", vertical="center")
    for c in range(4, 10):
        ws.cell(row=4, column=c).border = thin_border
    ws.row_dimensions[4].height = 20

    # Top Block Header in Row 4 for each individual phase (sc=11+p_idx*7, spacer at sc-1 is blank)
    for p_idx, ph in enumerate(phases):
        sc = 11 + p_idx * 7
        ec = sc + 5
        ws.merge_cells(start_row=4, start_column=sc, end_row=4, end_column=ec)
        cell_4 = ws.cell(row=4, column=sc, value=f"{ph['name']} ({ph.get('from_year', 0)}th to {ph.get('to_year', 0)}th yr)")
        cell_4.font = header_font
        cell_4.fill = header_fill
        cell_4.alignment = Alignment(horizontal="center", vertical="center")
        for c in range(sc, ec + 1):
            ws.cell(row=4, column=c).border = thin_border

    # Sl. No. header (A4:A5 merged, A6="A")
    ws.merge_cells("A4:A5")
    ws["A4"] = "Sl. No."
    ws["A4"].font = sub_font
    ws["A4"].fill = sub_fill
    ws["A4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A6"] = "A"
    ws["A6"].font = bold_font
    ws["A6"].fill = sub_fill
    ws["A6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=1).border = thin_border
    ws.cell(row=5, column=1).border = thin_border
    ws.cell(row=6, column=1).border = thin_border

    # Description header (B4:B5 merged, B6="MHS")
    ws.merge_cells("B4:B5")
    ws["B4"] = "Description"
    ws["B4"].font = sub_font
    ws["B4"].fill = sub_fill
    ws["B4"].alignment = Alignment(horizontal="center", vertical="center")
    ws["B6"] = "MHS"
    ws["B6"].font = bold_font
    ws["B6"].fill = sub_fill
    ws["B6"].alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=2).border = thin_border
    ws.cell(row=5, column=2).border = thin_border
    ws.cell(row=6, column=2).border = thin_border

    # Capacity header (C4:C6 merged)
    ws.merge_cells("C4:C6")
    ws["C4"] = "Capacity"
    ws["C4"].font = sub_font
    ws["C4"].fill = sub_fill
    ws["C4"].alignment = Alignment(horizontal="center", vertical="center")
    ws.cell(row=4, column=3).border = thin_border
    ws.cell(row=5, column=3).border = thin_border
    ws.cell(row=6, column=3).border = thin_border

    # Helper for rendering 6-column header block at any starting column
    def render_6col_header(sc):
        ws.merge_cells(start_row=5, start_column=sc, end_row=6, end_column=sc)
        c_ur = ws.cell(row=5, column=sc, value="Unit Rate")
        c_ur.font = sub_font; c_ur.fill = sub_fill; c_ur.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=sc).border = thin_border; ws.cell(row=6, column=sc).border = thin_border

        ws.merge_cells(start_row=5, start_column=sc+1, end_row=6, end_column=sc+1)
        c_qty = ws.cell(row=5, column=sc+1, value="Qty.")
        c_qty.font = sub_font; c_qty.fill = sub_fill; c_qty.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=sc+1).border = thin_border; ws.cell(row=6, column=sc+1).border = thin_border

        ws.merge_cells(start_row=5, start_column=sc+2, end_row=6, end_column=sc+2)
        c_uom = ws.cell(row=5, column=sc+2, value="UoM")
        c_uom.font = sub_font; c_uom.fill = sub_fill; c_uom.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=sc+2).border = thin_border; ws.cell(row=6, column=sc+2).border = thin_border

        ws.merge_cells(start_row=5, start_column=sc+3, end_row=6, end_column=sc+3)
        c_bc = ws.cell(row=5, column=sc+3, value="Base Cost")
        c_bc.font = sub_font; c_bc.fill = sub_fill; c_bc.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=sc+3).border = thin_border; ws.cell(row=6, column=sc+3).border = thin_border

        c_er5 = ws.cell(row=5, column=sc+4, value="Erection")
        c_er5.font = sub_font; c_er5.fill = sub_fill; c_er5.alignment = Alignment(horizontal="center", vertical="center")
        c_er5.border = thin_border
        c_er6 = ws.cell(row=6, column=sc+4, value=project.global_erection_pct)
        c_er6.number_format = '0%'
        c_er6.font = sub_font; c_er6.fill = sub_fill; c_er6.alignment = Alignment(horizontal="center", vertical="center")
        c_er6.border = thin_border

        ws.merge_cells(start_row=5, start_column=sc+5, end_row=6, end_column=sc+5)
        c_tc = ws.cell(row=5, column=sc+5, value="Total Cost")
        c_tc.font = sub_font; c_tc.fill = sub_fill; c_tc.alignment = Alignment(horizontal="center", vertical="center")
        ws.cell(row=5, column=sc+5).border = thin_border; ws.cell(row=6, column=sc+5).border = thin_border

    # Render All Phases column headers (sc=4) and each phase headers (sc=11+p_idx*7)
    render_6col_header(4)
    for p_idx in range(len(phases)):
        render_6col_header(11 + p_idx * 7)

    ws.row_dimensions[5].height = 20
    ws.row_dimensions[6].height = 20

    def write_group_banner(row_num, text):
        ws.merge_cells(start_row=row_num, start_column=1, end_row=row_num, end_column=9)
        ws.cell(row=row_num, column=1, value=text).font = bold_font
        for c in range(1, 10):
            ws.cell(row=row_num, column=c).border = thin_border
            ws.cell(row=row_num, column=c).fill = group_fill
        for p_idx in range(len(phases)):
            sc = 11 + p_idx * 7
            ws.merge_cells(start_row=row_num, start_column=sc, end_row=row_num, end_column=sc+5)
            ws.cell(row=row_num, column=sc, value=text).font = bold_font
            for c in range(sc, sc + 6):
                ws.cell(row=row_num, column=c).border = thin_border
                ws.cell(row=row_num, column=c).fill = group_fill
        ws.row_dimensions[row_num].height = 22

    # Banner 1: Belt Conveyor
    write_group_banner(7, "Belt Conveyor")

    def _get_item_uom_safe(item_list, default="Lot"):
        if not item_list:
            return default
        for item in item_list:
            if hasattr(item, 'unit_of_measure') and item.unit_of_measure:
                return item.unit_of_measure
            if item.category and item.category.name:
                return get_item_uom(item.category.name)
            if item.parent_category:
                return get_item_uom(item.parent_category)
        return default

    # Helper function to compute row cost across all phases and per-phase
    def compute_cost(filter_fn):
        matched = [i for i in items if filter_fn(i)]
        bc = round(sum(i.total_item_cost for i in matched), 2)
        ec = round(bc * project.global_erection_pct, 2)
        tc = round(bc + ec, 2)
        
        spec_text = "-"
        if matched:
            specs = []
            for mi in matched:
                if mi.selected_rate and isinstance(mi.selected_rate.specifications, dict):
                    s_vals = [str(v) for k, v in mi.selected_rate.specifications.items() if k != "Remarks" and str(v).strip()]
                    if s_vals: specs.append(" x ".join(s_vals))
                elif mi.category and mi.category.spec_schema:
                    specs.append(" | ".join([str(s) for s in mi.category.spec_schema if s != "Remarks"]))
            if specs:
                spec_text = " / ".join(list(dict.fromkeys(specs)))
            else:
                spec_text = "Selected"

        phase_costs = []
        for ph in phases:
            ph_matched = [i for i in matched if i.phase_name == ph["name"]]
            ph_bc = round(sum(i.total_item_cost for i in ph_matched), 2)
            ph_ec = round(ph_bc * project.global_erection_pct, 2)
            ph_tc = round(ph_bc + ph_ec, 2)
            ph_qty = round(sum(i.quantity for i in ph_matched), 2) if ph_matched else 0
            ph_uom = _get_item_uom_safe(ph_matched, default=_get_item_uom_safe(matched, default="Lot"))
            ph_ur = round(ph_bc / ph_qty, 2) if ph_qty > 0 else ph_bc
            phase_costs.append({
                "bc": ph_bc, "ec": ph_ec, "tc": ph_tc,
                "qty": ph_qty, "uom": ph_uom, "unit_rate": ph_ur
            })

        return bc, ec, tc, matched, spec_text, phase_costs

    def write_row_data(row_num, sl_str, desc_str, cap_str, unit_rate_val, qty_val, uom_val, bc_val, ec_val, tc_val, p_costs, has_margin=False):
        ws.cell(row=row_num, column=1, value=sl_str).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=2, value=desc_str)
        ws.cell(row=row_num, column=3, value=cap_str).alignment = Alignment(horizontal="center")
        ws.cell(row=row_num, column=5, value=qty_val if qty_val > 0 else "").alignment = Alignment(horizontal="right")
        ws.cell(row=row_num, column=6, value=uom_val).alignment = Alignment(horizontal="center")

        # Base Cost (Col G) is the exact calculated cost
        ws.cell(row=row_num, column=7, value=bc_val).number_format = '₹#,##0.00'

        # Unit Rate (Col D) is the formula =Base Cost / Qty
        if isinstance(qty_val, (int, float)) and qty_val > 0:
            ws.cell(row=row_num, column=4, value=f"=IF(E{row_num}>0, G{row_num}/E{row_num}, G{row_num})").number_format = '₹#,##0.00'
        else:
            ws.cell(row=row_num, column=4, value=unit_rate_val).number_format = '₹#,##0.00'

        if not has_margin and project.global_margin_pct > 0:
            ws.cell(row=row_num, column=8, value=f"=G{row_num}*(1+{project.global_margin_pct})*H$6").number_format = '₹#,##0.00'
        else:
            ws.cell(row=row_num, column=8, value=f"=G{row_num}*H$6").number_format = '₹#,##0.00'

        if not has_margin and project.global_margin_pct > 0:
            ws.cell(row=row_num, column=9, value=f"=G{row_num}*(1+{project.global_margin_pct})+H{row_num}").number_format = '₹#,##0.00'
        else:
            ws.cell(row=row_num, column=9, value=f"=G{row_num}+H{row_num}").number_format = '₹#,##0.00'

        for c in range(1, 10):
            ws.cell(row=row_num, column=c).border = thin_border

        for p_idx, pc in enumerate(p_costs):
            sc = 11 + p_idx * 7
            sc_ur = get_column_letter(sc)
            sc_qty = get_column_letter(sc+1)
            sc_uom = get_column_letter(sc+2)
            sc_bc = get_column_letter(sc+3)
            sc_ec = get_column_letter(sc+4)
            sc_tc = get_column_letter(sc+5)

            ws.cell(row=row_num, column=sc+1, value=pc["qty"] if pc["qty"] > 0 else "").alignment = Alignment(horizontal="right")
            ws.cell(row=row_num, column=sc+2, value=pc["uom"]).alignment = Alignment(horizontal="center")
            ws.cell(row=row_num, column=sc+3, value=pc["bc"]).number_format = '₹#,##0.00'

            if isinstance(pc["qty"], (int, float)) and pc["qty"] > 0:
                ws.cell(row=row_num, column=sc, value=f"=IF({sc_qty}{row_num}>0, {sc_bc}{row_num}/{sc_qty}{row_num}, {sc_bc}{row_num})").number_format = '₹#,##0.00'
            else:
                ws.cell(row=row_num, column=sc, value=pc["unit_rate"]).number_format = '₹#,##0.00'

            if not has_margin and project.global_margin_pct > 0:
                ws.cell(row=row_num, column=sc+4, value=f"={sc_bc}{row_num}*(1+{project.global_margin_pct})*{sc_ec}$6").number_format = '₹#,##0.00'
            else:
                ws.cell(row=row_num, column=sc+4, value=f"={sc_bc}{row_num}*{sc_ec}$6").number_format = '₹#,##0.00'

            if not has_margin and project.global_margin_pct > 0:
                ws.cell(row=row_num, column=sc+5, value=f"={sc_bc}{row_num}*(1+{project.global_margin_pct})+{sc_ec}{row_num}").number_format = '₹#,##0.00'
            else:
                ws.cell(row=row_num, column=sc+5, value=f"={sc_bc}{row_num}+{sc_ec}{row_num}").number_format = '₹#,##0.00'

            for c in range(sc, sc + 6):
                ws.cell(row=row_num, column=c).border = thin_border

        ws.row_dimensions[row_num].height = 20

    # Row 8: Sl 1 - Supply (Belt conv.+Tech. strls+chutes)
    bc_1, ec_1, tc_1, m_1, _, pc_1 = compute_cost(lambda i: i.parent_category == "Belt Conveyor")
    belt_items = [i for i in m_1 if i.category and 'belt' in i.category.name.lower() and 'cleaner' not in i.category.name.lower() and 'weigher' not in i.category.name.lower()]
    
    if project.conveyor_length_mtr and project.conveyor_length_mtr > 0:
        qty_1 = round(project.conveyor_length_mtr, 2)
        uom_1 = "R.Mtr."
    elif belt_items:
        qty_1 = round(sum(i.quantity for i in belt_items), 2)
        uom_1 = _get_item_uom_safe(belt_items, default="R.Mtr.")
    else:
        qty_1 = 0 if not m_1 else 1
        uom_1 = "R.Mtr."

    bc_1_with_margin = round(sum(i.total_item_cost for i in m_1), 2)
    unit_rate_1 = round(bc_1_with_margin / qty_1, 2) if qty_1 > 0 else bc_1_with_margin

    for pc_item, ph_dict in zip(pc_1, phases):
        ph_matched = [i for i in m_1 if i.phase_name == ph_dict["name"]]
        ph_bc_margin = round(sum(i.total_item_cost for i in ph_matched), 2)
        ph_qty = project.conveyor_length_mtr if (project.conveyor_length_mtr and project.conveyor_length_mtr > 0 and ph_matched) else pc_item["qty"]
        pc_item["bc"] = ph_bc_margin
        pc_item["qty"] = ph_qty
        pc_item["uom"] = "R.Mtr." if (project.conveyor_length_mtr and project.conveyor_length_mtr > 0) else pc_item["uom"]
        pc_item["unit_rate"] = round(ph_bc_margin / ph_qty, 2) if ph_qty > 0 else ph_bc_margin

    write_row_data(8, "1", "Supply (Belt conv.+Tech. strls+chutes)", "-", unit_rate_1, qty_1, uom_1, bc_1_with_margin, ec_1, tc_1, pc_1, has_margin=True)

    # Row 9: Sl 2 - Auxiliary Equipments
    bc_2, ec_2, tc_2, m_2, _, pc_2 = compute_cost(lambda i: i.parent_category == "Auxiliary Equipment")
    write_row_data(9, "2", "Auxiliary Equipments (Divertor Gate + BW + ILMS + MD + Sampler + Hoist)", "-", bc_2, 1 if m_2 else 0, "Lot", bc_2, ec_2, tc_2, pc_2)

    # Row 10: Sl 3 - Hopper Above Crusher 1 x 500 T
    bc_3, ec_3, tc_3, m_3, spec_3, pc_3 = compute_cost(lambda i: i.parent_category == "Hopper Above Crusher")
    write_row_data(10, "3", "Hopper Above Crusher 1 x 500 T", spec_3, bc_3, 1 if m_3 else 0, "Lot", bc_3, ec_3, tc_3, pc_3)

    # Banner 2: Major Equipments - Supply
    write_group_banner(11, "Major Equipments - Supply")

    major_defs = [
        (4, "Sizer (Secondary)", lambda i: i.category and "sizer" in i.category.name.lower()),
        (5, "Vibrating Feeder", lambda i: i.category and "vibrating" in i.category.name.lower()),
        (6, "Truck loading system + (2 x 150T)", lambda i: i.category and "truck" in i.category.name.lower()),
        (7, "Rapid Wagon loading system (TLO)", lambda i: i.category and "rapid" in i.category.name.lower()),
        (8, "Stacker cum Reclaimer", lambda i: i.category and "stacker" in i.category.name.lower()),
        (9, "IPC (Skid Mounted System incl. 4 x 250T Hopper, Apron Feeder, Sizers etc.)", lambda i: i.category and "ipc" in i.category.name.lower() and "shifting" not in i.category.name.lower()),
        (10, "Shifting of IPC system", lambda i: i.category and "shifting" in i.category.name.lower()),
        (11, "Workshop", lambda i: i.category and "workshop" in i.category.name.lower()),
    ]

    row_idx = 12
    total_mech_base = bc_1 + bc_2 + bc_3
    total_mech_erection = ec_1 + ec_2 + ec_3
    total_mech_total = tc_1 + tc_2 + tc_3
    
    total_phase_base = [pc_1[idx]["bc"] + pc_2[idx]["bc"] + pc_3[idx]["bc"] for idx in range(len(phases))]
    total_phase_erection = [pc_1[idx]["ec"] + pc_2[idx]["ec"] + pc_3[idx]["ec"] for idx in range(len(phases))]
    total_phase_total = [pc_1[idx]["tc"] + pc_2[idx]["tc"] + pc_3[idx]["tc"] for idx in range(len(phases))]

    matched_major_ids = set()

    for sl, desc, fn in major_defs:
        bc_m, ec_m, tc_m, matched, spec_m, pc_m = compute_cost(fn)
        for mi in matched: matched_major_ids.add(mi.id)
        total_mech_base += bc_m
        total_mech_erection += ec_m
        total_mech_total += tc_m
        for p_idx in range(len(phases)):
            total_phase_base[p_idx] += pc_m[p_idx]["bc"]
            total_phase_erection[p_idx] += pc_m[p_idx]["ec"]
            total_phase_total[p_idx] += pc_m[p_idx]["tc"]

        qty_m = round(sum(i.quantity for i in matched), 2) if matched else 0
        uom_m = _get_item_uom_safe(matched, default=("Nos." if sl in [4, 5, 6, 7, 8] else "Lot"))
        unit_rate_m = round(bc_m / qty_m, 2) if qty_m > 0 else bc_m
        write_row_data(row_idx, str(sl), desc, spec_m, unit_rate_m, qty_m, uom_m, bc_m, ec_m, tc_m, pc_m)
        row_idx += 1

    # Catch any remaining items under Major Equipment
    remaining_major = [i for i in items if i.parent_category == "Major Equipment" and i.id not in matched_major_ids]
    if remaining_major:
        subcat_groups = {}
        for item in remaining_major:
            c_name = item.category.name if item.category else f"Item #{item.id}"
            subcat_groups.setdefault(c_name, []).append(item)
        for sl_offset, (s_name, s_items) in enumerate(subcat_groups.items(), 12):
            s_ids = set(i.id for i in s_items)
            bc_rm, ec_rm, tc_rm, _, spec_rm, pc_rm = compute_cost(lambda i: i.id in s_ids)
            total_mech_base += bc_rm
            total_mech_erection += ec_rm
            total_mech_total += tc_rm
            for p_idx in range(len(phases)):
                total_phase_base[p_idx] += pc_rm[p_idx]["bc"]
                total_phase_erection[p_idx] += pc_rm[p_idx]["ec"]
                total_phase_total[p_idx] += pc_rm[p_idx]["tc"]

            qty_rm = round(sum(i.quantity for i in s_items), 2)
            uom_rm = _get_item_uom_safe(s_items, default="Lot")
            unit_rate_rm = round(bc_rm / qty_rm, 2) if qty_rm > 0 else bc_rm
            write_row_data(row_idx, str(sl_offset), s_name, spec_rm, unit_rate_rm, qty_rm, uom_rm, bc_rm, ec_rm, tc_rm, pc_rm)
            row_idx += 1

    double_bottom_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='double', color='000000')
    )

    def write_subtotal_row(row_num, start_row, end_row):
        ws.merge_cells(f"B{row_num}:F{row_num}")
        ws.cell(row=row_num, column=2, value="TOTAL COST>>>").font = bold_font
        ws.cell(row=row_num, column=2).alignment = Alignment(horizontal="right", vertical="center")
        for c in range(1, 10):
            ws.cell(row=row_num, column=c).fill = subtotal_fill
            ws.cell(row=row_num, column=c).border = double_bottom_border

        ws.cell(row=row_num, column=7, value=f"=SUM(G{start_row}:G{end_row})").number_format = '₹#,##0.00'
        ws.cell(row=row_num, column=7).font = bold_font
        ws.cell(row=row_num, column=8, value=f"=SUM(H{start_row}:H{end_row})").number_format = '₹#,##0.00'
        ws.cell(row=row_num, column=8).font = bold_font
        ws.cell(row=row_num, column=9, value=f"=SUM(I{start_row}:I{end_row})").number_format = '₹#,##0.00'
        ws.cell(row=row_num, column=9).font = bold_font

        for p_idx in range(len(phases)):
            sc = 11 + p_idx * 7
            ws.merge_cells(start_row=row_num, start_column=sc, end_row=row_num, end_column=sc+2)
            ws.cell(row=row_num, column=sc, value="TOTAL COST>>>").font = bold_font
            ws.cell(row=row_num, column=sc).alignment = Alignment(horizontal="right", vertical="center")
            for c in range(sc, sc + 6):
                ws.cell(row=row_num, column=c).fill = subtotal_fill
                ws.cell(row=row_num, column=c).border = double_bottom_border

            sc_bc = get_column_letter(sc+3)
            sc_ec = get_column_letter(sc+4)
            sc_tc = get_column_letter(sc+5)

            ws.cell(row=row_num, column=sc+3, value=f"=SUM({sc_bc}{start_row}:{sc_bc}{end_row})").number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=sc+3).font = bold_font
            ws.cell(row=row_num, column=sc+4, value=f"=SUM({sc_ec}{start_row}:{sc_ec}{end_row})").number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=sc+4).font = bold_font
            ws.cell(row=row_num, column=sc+5, value=f"=SUM({sc_tc}{start_row}:{sc_tc}{end_row})").number_format = '₹#,##0.00'
            ws.cell(row=row_num, column=sc+5).font = bold_font

        ws.row_dimensions[row_num].height = 22

    # TOTAL COST>>> Row for Mechanical
    mech_end_row = row_idx - 1
    write_subtotal_row(row_idx, start_row=8, end_row=mech_end_row)
    row_idx += 1

    # Banner 3: Utility
    write_group_banner(row_idx, "Utility")
    row_idx += 1
    util_start_row = row_idx

    bc_util, ec_util, tc_util, m_util, spec_util, pc_util = compute_cost(lambda i: str(i.domain.value if hasattr(i.domain, 'value') else i.domain).lower() == "utility" or (i.category and 'util' in i.category.name.lower()))
    write_row_data(row_idx, "1", "Utilities", spec_util, bc_util, 1 if m_util else 0, "Lot", bc_util, ec_util, tc_util, pc_util)
    util_end_row = row_idx
    row_idx += 1

    # TOTAL COST>>> Row for Utility
    write_subtotal_row(row_idx, start_row=util_start_row, end_row=util_end_row)

    adjust_column_widths(ws)
    return total_mech_base + bc_util

def populate_mech_detailed_sheet(ws, project, line_items, phase_name=None, phase_info=None):
    if phase_name and phase_name != "All":
        ws.title = str(phase_name)[:31]
    else:
        ws.title = "Mech Detailed"
    title_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    header_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
    bold_font = Font(name="Calibri", size=10, bold=True, color="0F172A")
    header_fill = PatternFill(start_color="4F6D7A", end_color="4F6D7A", fill_type="solid") # Soft steel blue
    sub_header_fill = PatternFill(start_color="E2E8F0", end_color="E2E8F0", fill_type="solid") # Soft cool grey
    thin_border = Border(
        left=Side(style='thin', color='CBD5E1'),
        right=Side(style='thin', color='CBD5E1'),
        top=Side(style='thin', color='CBD5E1'),
        bottom=Side(style='thin', color='CBD5E1')
    )

    items = [i for i in line_items if str(i.domain.value if hasattr(i.domain, 'value') else i.domain).lower() == "mechanical" or (i.category and str(i.category.domain.value if hasattr(i.category.domain, 'value') else i.category.domain).lower() == "mechanical")]
    if phase_name and phase_name != "All":
        items = [i for i in items if i.phase_name == phase_name]

    ws.merge_cells("A1:I1")
    if phase_info:
        ws["A1"] = f"DETAILED MECHANICAL SPECIFICATION REPORT: {project.name.upper()} ({phase_name.upper()} - {phase_info.get('from_year', 0)}th to {phase_info.get('to_year', 0)}th yr)"
    elif phase_name and phase_name != "All":
        ws["A1"] = f"DETAILED MECHANICAL SPECIFICATION REPORT: {project.name.upper()} ({phase_name.upper()})"
    else:
        ws["A1"] = f"DETAILED MECHANICAL SPECIFICATION REPORT: {project.name.upper()} (BY EQUIPMENT GROUP)"
    ws["A1"].font = title_font
    ws["A1"].fill = header_fill
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    headers = [
        "Sl. No.",
        "Description",
        "Specification",
        "Belt Width",
        "Vendor Name",
        "Unit Rate (INR)",
        "Qty.",
        "UoM",
        "Base Cost (INR)",
        "EPC Markup",
        "Total Cost (INR)"
    ]
    for col_num, h_text in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col_num, value=h_text)
        cell.font = bold_font
        cell.fill = sub_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    ws.row_dimensions[3].height = 24

    # Group items by equipment category name (sub-heading)
    grouped_items = {}
    for it in items:
        cname = it.category.name if it.category else f"Category #{it.category_id}"
        if cname not in grouped_items:
            grouped_items[cname] = []
        grouped_items[cname].append(it)

    if not grouped_items:
        ws.merge_cells("A4:K4")
        msg = f"No mechanical line items added for {phase_name} yet." if phase_name and phase_name != "All" else "No mechanical line items added yet. Please select categories above."
        ws.cell(row=4, column=1, value=msg).alignment = Alignment(horizontal="center")
        ws.row_dimensions[4].height = 25
        return

    row_idx = 4
    for group_idx, (cat_name, cat_items) in enumerate(grouped_items.items(), 1):
        # Sub-heading banner row for similar equipment clubbed together
        ws.merge_cells(f"A{row_idx}:K{row_idx}")
        cat_schema = cat_items[0].category.spec_schema if (cat_items and cat_items[0].category and cat_items[0].category.spec_schema) else []
        
        # Filter schema to remove Remarks, Type columns, and Belt Width
        filtered_schema = [
            s for s in cat_schema 
            if str(s).strip().lower() not in ["remarks", "remark"] 
            and "type" not in str(s).strip().lower() 
            and str(s).strip().lower() not in ["belt width", "bw", "bw (mm)", "belt_width"]
        ]
        if filtered_schema:
            schema_str = " x ".join([str(s) for s in filtered_schema])
            banner_text = f"{cat_name} ({schema_str})"
        else:
            banner_text = f"{cat_name}"

        ws.cell(row=row_idx, column=1, value=banner_text).font = Font(name="Calibri", size=11, bold=True, color="1e293b")
        ws.cell(row=row_idx, column=1).fill = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
        for c in range(1, 12):
            ws.cell(row=row_idx, column=c).border = thin_border
        ws.row_dimensions[row_idx].height = 22
        row_idx += 1

        for it_idx, it in enumerate(cat_items, 1):
            sl_no = f"{group_idx}.{it_idx}"
            spec_dict = it.selected_rate.specifications if it.selected_rate and it.selected_rate.specifications else {}
            
            # 1. Description: show equipment type if category has_type (e.g. Drive Pulley / Tail Pulley), else subcategory name
            has_type = (it.category and it.category.has_type) or any('type' in str(k).lower() for k in spec_dict.keys())
            desc_val = cat_name
            if has_type:
                type_val = None
                for k, v in spec_dict.items():
                    if 'type' in str(k).lower() and v is not None and str(v).strip() != "":
                        type_val = str(v).strip()
                        break
                if not type_val and 'pulley' in cat_name.lower():
                    rem_str = str(spec_dict.get('Remarks', '') or (it.selected_rate.remarks if it.selected_rate else '')).lower()
                    if 'drive pulley' in rem_str:
                        type_val = "Drive Pulley"
                    elif 'tail pulley' in rem_str:
                        type_val = "Tail Pulley"
                    elif 'bend pulley' in rem_str:
                        type_val = "Bend Pulley"
                    elif 'snub pulley' in rem_str:
                        type_val = "Snub Pulley"
                    elif 'head pulley' in rem_str:
                        type_val = "Head Pulley"
                if type_val:
                    desc_val = type_val

            # 2. Specification: Exact Specification in 'X' Strings Format excluding Remarks, Type, and Belt Width
            x_string = ""
            if isinstance(spec_dict, dict) and spec_dict:
                x_keys = [k for k in spec_dict.keys() if 'x' in str(k).lower() and not any(w in str(k).lower() for w in ['remarks', 'belt width', 'bw'])]
                if x_keys and spec_dict[x_keys[0]] is not None and str(spec_dict[x_keys[0]]).strip() != "":
                    x_string = str(spec_dict[x_keys[0]]).strip()
                else:
                    vals = [
                        str(v).strip() for k, v in spec_dict.items() 
                        if str(k).strip().lower() not in ["remarks", "remark"] 
                        and "type" not in str(k).strip().lower() 
                        and str(k).strip().lower() not in ["belt width", "bw", "bw (mm)", "belt_width"]
                        and v is not None and str(v).strip() != ""
                    ]
                    x_string = " x ".join(vals)
            else:
                x_string = "Standard Specification"

            # 3. Belt Width column: separate column, show if present else blank
            has_bw = (it.category and it.category.has_bw) or any(str(k).lower() in ['belt width', 'bw', 'bw (mm)', 'belt_width'] for k in spec_dict.keys())
            bw_val = ""
            if has_bw:
                for k, v in spec_dict.items():
                    if str(k).lower() in ['belt width', 'bw', 'bw (mm)', 'belt_width'] and v is not None and str(v).strip() != "":
                        bw_val = str(v).strip()
                        break

            vendor = it.selected_rate.vendor_name if it.selected_rate else "Standard Vendor"
            unit_rate = it.calculated_escalated_rate
            qty = it.quantity
            uom = get_item_uom(cat_name)
            item_margin = it.selected_rate.margin_pct if (it.selected_rate and hasattr(it.selected_rate, 'margin_pct') and it.selected_rate.margin_pct is not None) else project.global_margin_pct

            ws.cell(row=row_idx, column=1, value=sl_no).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=2, value=desc_val)
            ws.cell(row=row_idx, column=3, value=x_string).font = Font(name="Calibri", size=10, bold=True, color="0f172a")
            ws.cell(row=row_idx, column=4, value=bw_val).alignment = Alignment(horizontal="center")
            ws.cell(row=row_idx, column=5, value=vendor)

            ur_cell = ws.cell(row=row_idx, column=6, value=unit_rate)
            ur_cell.number_format = '₹#,##0.00'
            ur_cell.alignment = Alignment(horizontal="right")

            ws.cell(row=row_idx, column=7, value=qty).alignment = Alignment(horizontal="right")
            ws.cell(row=row_idx, column=8, value=uom).alignment = Alignment(horizontal="center")

            # Col 9 (I): Base Cost = Unit Rate * Qty
            if isinstance(qty, (int, float)) and qty > 0 and isinstance(unit_rate, (int, float)) and unit_rate > 0:
                bc_cell = ws.cell(row=row_idx, column=9, value=f"=F{row_idx}*G{row_idx}")
            else:
                bc_cell = ws.cell(row=row_idx, column=9, value=round(unit_rate * qty, 2))
            bc_cell.number_format = '₹#,##0.00'
            bc_cell.alignment = Alignment(horizontal="right")

            # Col 10 (J): EPC Markup
            margin_cell = ws.cell(row=row_idx, column=10, value=item_margin)
            margin_cell.number_format = '0.0%'
            margin_cell.alignment = Alignment(horizontal="center")

            # Col 11 (K): Total Cost = Base Cost * (1 + EPC Markup)
            if isinstance(qty, (int, float)) and qty > 0 and isinstance(unit_rate, (int, float)) and unit_rate > 0:
                tc_cell = ws.cell(row=row_idx, column=11, value=f"=I{row_idx}*(1+J{row_idx})")
            else:
                tc_cell = ws.cell(row=row_idx, column=11, value=round(unit_rate * qty * (1.0 + item_margin), 2))
            tc_cell.number_format = '₹#,##0.00'
            tc_cell.alignment = Alignment(horizontal="right")

            for c in range(1, 12):
                ws.cell(row=row_idx, column=c).border = thin_border
            ws.row_dimensions[row_idx].height = 19
            row_idx += 1

    # Add TOTAL COST subtotal row with SUM formula at the bottom
    ws.merge_cells(f"B{row_idx}:H{row_idx}")
    ws.cell(row=row_idx, column=2, value="TOTAL COST>>>").font = bold_font
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal="right", vertical="center")
    for c in range(1, 12):
        ws.cell(row=row_idx, column=c).fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
        ws.cell(row=row_idx, column=c).border = Border(left=Side(style='thin', color='CBD5E1'), right=Side(style='thin', color='CBD5E1'), top=Side(style='thin', color='CBD5E1'), bottom=Side(style='double', color='000000'))

    ws.cell(row=row_idx, column=9, value=f"=SUM(I4:I{row_idx-1})").number_format = '₹#,##0.00'
    ws.cell(row=row_idx, column=9).font = bold_font

    ws.cell(row=row_idx, column=11, value=f"=SUM(K4:K{row_idx-1})").number_format = '₹#,##0.00'
    ws.cell(row=row_idx, column=11).font = bold_font
    ws.row_dimensions[row_idx].height = 22
    row_idx += 1

    adjust_column_widths(ws)

@router.get("/{project_id}/export-excel")
def export_project_excel(
    project_id: int,
    domain: Optional[str] = None,
    db: Session = Depends(get_db),
    current_user: models.User = Depends(auth.get_current_user)
):
    project = db.query(models.Project).filter(models.Project.id == project_id).first()
    if not project:
        raise HTTPException(status_code=404, detail="Project not found")
    
    line_items = db.query(models.EstimateLineItem).filter(models.EstimateLineItem.project_id == project_id).all()
    wb = openpyxl.Workbook()
    phases = project.phases or []

    if domain == "Mechanical":
        ws_sum = wb.active
        populate_mech_summary_sheet(ws_sum, project, line_items)
        if phases:
            for ph in phases:
                ws_det = wb.create_sheet(title=str(ph["name"])[:31])
                populate_mech_detailed_sheet(ws_det, project, line_items, phase_name=ph["name"], phase_info=ph)
        else:
            ws_det = wb.create_sheet(title="Mech Detailed")
            populate_mech_detailed_sheet(ws_det, project, line_items, phase_name="All", phase_info=None)
        filename = f"{project.name.replace(' ', '_')}_Mechanical_Report.xlsx"
    elif domain in ["Electrical", "Civil"]:
        ws = wb.active
        populate_sheet(ws, project, line_items, domain)
        filename = f"{domain}_Report.xlsx"
    else:
        # Full project report with Mechanical Summary + Detailed phase sheets, plus Electrical and Civil
        ws_mech_sum = wb.active
        populate_mech_summary_sheet(ws_mech_sum, project, line_items)
        
        if phases:
            for ph in phases:
                ws_det = wb.create_sheet(title=str(ph["name"])[:31])
                populate_mech_detailed_sheet(ws_det, project, line_items, phase_name=ph["name"], phase_info=ph)
        else:
            ws_mech_det = wb.create_sheet(title="Mech Detailed")
            populate_mech_detailed_sheet(ws_mech_det, project, line_items, phase_name="All", phase_info=None)
        
        ws_elec = wb.create_sheet(title="Electrical")
        elec_sub = populate_sheet(ws_elec, project, line_items, "Electrical")
        
        ws_civil = wb.create_sheet(title="Civil")
        civil_sub = populate_sheet(ws_civil, project, line_items, "Civil")
        
        # Sheet 5: Executive Summary
        ws_sum = wb.create_sheet(title="Executive Summary")
        title_font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="0f172a", end_color="0f172a", fill_type="solid")
        bold_font = Font(name="Calibri", size=11, bold=True)
        thin_border = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )
        
        ws_sum.merge_cells("A1:F1")
        ws_sum["A1"] = f"EXECUTIVE PROJECT ESTIMATION SUMMARY: {project.name.upper()}"
        ws_sum["A1"].font = title_font
        ws_sum["A1"].fill = header_fill
        ws_sum["A1"].alignment = Alignment(horizontal="center", vertical="center")
        ws_sum.row_dimensions[1].height = 35
        
        ws_sum["A3"] = "Client Name:"
        ws_sum["B3"] = project.client
        ws_sum["A4"] = "Date Generated:"
        ws_sum["B4"] = datetime.now().strftime("%d/%m/%Y")
        ws_sum["D3"] = "Global EPC Markup:"
        ws_sum["E3"] = f"{project.global_margin_pct * 100:.1f}%"
        ws_sum["D4"] = "Global Erection:"
        ws_sum["E4"] = f"{project.global_erection_pct * 100:.1f}%"
        for cell in ["A3", "A4", "D3", "D4", "B3", "B4", "E3", "E4"]:
            ws_sum[cell].font = bold_font

        headers = ["Discipline / Category", "Line Items Count", "Total Equipment Cost (INR)"]
        for col_num, h_text in enumerate(headers, 2):
            cell = ws_sum.cell(row=6, column=col_num, value=h_text)
            cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="1e293b", end_color="1e293b", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border
        ws_sum.row_dimensions[6].height = 24
        
        mech_items = [i for i in line_items if i.domain == models.DomainType.Mechanical or (i.category and i.category.domain == models.DomainType.Mechanical)]
        mech_sub = sum(i.total_item_cost for i in mech_items)
        mech_count = len(mech_items)
        elec_count = len([i for i in line_items if i.domain == models.DomainType.Electrical or (i.category and i.category.domain == models.DomainType.Electrical)])
        civil_count = len([i for i in line_items if i.domain == models.DomainType.Civil or (i.category and i.category.domain == models.DomainType.Civil)])
        
        subtotal = mech_sub + elec_sub + civil_sub
        margin_cost = round(subtotal * project.global_margin_pct, 2)
        erection_cost = round((subtotal + margin_cost) * project.global_erection_pct, 2)
        grand_total = round(subtotal + margin_cost + erection_cost, 2)
        
        rows_data = [
            ("Mechanical Discipline", mech_count, mech_sub),
            ("Electrical Discipline", elec_count, elec_sub),
            ("Civil Discipline", civil_count, civil_sub),
            ("EQUIPMENT SUBTOTAL", mech_count + elec_count + civil_count, subtotal),
            (f"EPC Markup Amount ({project.global_margin_pct * 100:.1f}%)", "-", margin_cost),
            (f"Erection Amount ({project.global_erection_pct * 100:.1f}%)", "-", erection_cost),
            ("GRAND ESTIMATE TOTAL", "-", grand_total)
        ]
        
        r_idx = 7
        for label, count, val in rows_data:
            ws_sum.cell(row=r_idx, column=2, value=label).font = bold_font
            ws_sum.cell(row=r_idx, column=3, value=count).alignment = Alignment(horizontal="center")
            val_cell = ws_sum.cell(row=r_idx, column=4, value=val)
            val_cell.number_format = '₹#,##0.00'
            val_cell.alignment = Alignment(horizontal="right")
            val_cell.font = bold_font if "TOTAL" in label or "SUBTOTAL" in label else Font(name="Calibri", size=11)
            if "GRAND" in label:
                val_cell.font = Font(name="Calibri", size=13, bold=True, color="004B23")
            for c in range(2, 5):
                ws_sum.cell(row=r_idx, column=c).border = thin_border
            ws_sum.row_dimensions[r_idx].height = 22 if "TOTAL" in label else 19
            r_idx += 1
            
        for col in ws_sum.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws_sum.column_dimensions[col_letter].width = max(max_len + 4, 16)
            
        clean_name = "".join(x for x in project.name if x.isalnum() or x in " _-").replace(" ", "_")
        filename = f"{clean_name}_Full_Project_Report.xlsx"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'}
    )
